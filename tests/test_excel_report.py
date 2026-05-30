from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from dbc_compare_tool.core.models import Change, ComparisonResult
from dbc_compare_tool.report.excel import write_excel_report


class ExcelReportTests(unittest.TestCase):
    def test_report_contains_exact_required_sheets(self) -> None:
        result = ComparisonResult(
            message_changes=[
                Change("Bus_A.dbc", "Renamed", "BCM_Status", "Body_Status", 0.95, "CAN ID matched", 100)
            ],
            signal_changes=[
                Change(
                    dbc_file="Bus_A.dbc",
                    parent_message="Body_Status",
                    change_type="Renamed",
                    old_name="VehicleSpeed",
                    new_name="VehSpd",
                    confidence=0.97,
                    description="Start bit matched",
                )
            ],
        )
        with tempfile.TemporaryDirectory() as temp:
            report_path = Path(temp) / "report.xlsx"
            write_excel_report(result, report_path)

            workbook = load_workbook(report_path)

        self.assertEqual(workbook.sheetnames, ["Summary", "Message Details", "Signal Details"])
        summary_rows = {row[0]: row[1] for row in workbook["Summary"].iter_rows(min_row=2, values_only=True)}
        self.assertEqual(summary_rows["Messages Renamed"], 1)
        self.assertEqual(summary_rows["Signals Renamed"], 1)
        self.assertEqual(summary_rows["Total Changes"], 2)


if __name__ == "__main__":
    unittest.main()

