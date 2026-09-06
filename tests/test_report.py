"""Excel report tests.

The writer is exercised end to end — written to disk, reopened with openpyxl,
and read back. Without this the report layer can crash or drop a sheet while
the whole suite stays green, which is exactly what a release build would then
ship.
"""

import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from dbc_compare_tool.core.comparator import DbcComparator
from dbc_compare_tool.report.excel import write_excel_report

_EXAMPLES = Path(__file__).resolve().parents[1] / "examples"
_OLD_FOLDER = _EXAMPLES / "old"
_NEW_FOLDER = _EXAMPLES / "new"


def _rows(sheet) -> list[list]:
    return [list(row) for row in sheet.iter_rows(min_row=2, values_only=True)]


class BaselineReportTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.result = DbcComparator().compare_folders(_OLD_FOLDER, _NEW_FOLDER)

    def setUp(self):
        self._tmp = tempfile.TemporaryDirectory()
        self.addCleanup(self._tmp.cleanup)
        self.output = Path(self._tmp.name) / "report.xlsx"
        write_excel_report(self.result, self.output)
        self.workbook = load_workbook(self.output)
        self.addCleanup(self.workbook.close)

    def test_workbook_has_the_five_documented_sheets_in_order(self):
        self.assertEqual(
            self.workbook.sheetnames,
            ["Summary", "DBC Overview", "Message Details", "Signal Details", "Property Diff"],
        )

    def test_summary_total_matches_the_result(self):
        sheet = self.workbook["Summary"]
        totals = {row[0]: row[1] for row in sheet.iter_rows(min_row=4, values_only=True)}
        self.assertEqual(totals["Total Changes"], self.result.summary()["Total Changes"])
        self.assertEqual(totals["Messages Renamed"], self.result.summary()["Messages Renamed"])

    def test_every_signal_change_reaches_the_signal_sheet(self):
        rows = _rows(self.workbook["Signal Details"])
        self.assertEqual(len(rows), len(self.result.signal_changes))
        renamed = {(row[3], row[4]) for row in rows if row[2] == "Renamed"}
        self.assertIn(("VehicleSpeed", "VehSpd"), renamed)

    def test_can_id_is_written_in_hexadecimal(self):
        rows = _rows(self.workbook["Message Details"])
        self.assertTrue(rows)
        self.assertTrue(all(str(row[4]).startswith("0x") for row in rows if row[4]))

    def test_property_diff_carries_old_and_new_values(self):
        rows = _rows(self.workbook["Property Diff"])
        self.assertTrue(rows)
        factor_rows = [row for row in rows if row[6] == "Factor"]
        self.assertTrue(factor_rows, "the EngineRPM factor change should appear")
        self.assertNotEqual(factor_rows[0][7], factor_rows[0][8])

    def test_overview_lists_every_file_pair(self):
        rows = _rows(self.workbook["DBC Overview"])
        self.assertEqual(len(rows), len(self.result.file_pairs))
        self.assertEqual(rows[0][1], "Matched")

    def test_missing_parent_directory_is_created(self):
        nested = Path(self._tmp.name) / "a" / "b" / "report.xlsx"
        write_excel_report(self.result, nested)
        self.assertTrue(nested.is_file())


if __name__ == "__main__":
    unittest.main()
