"""Shared Excel styling so every generated report looks like one product."""

from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADER_BG = "2E74B5"   # modern blue header
HEADER_FG = "FFFFFF"

THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def format_header_row(sheet, row: int = 1) -> None:
    fill = PatternFill("solid", fgColor=HEADER_BG)
    font = Font(color=HEADER_FG, bold=True, size=11)
    for cell in sheet[row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    sheet.row_dimensions[row].height = 26


def autosize_columns(sheet, cap: int = 60, minimum: int = 12) -> None:
    for col_cells in sheet.columns:
        max_len = max(len(str(cell.value or "")) for cell in col_cells)
        col_letter = get_column_letter(col_cells[0].column)
        sheet.column_dimensions[col_letter].width = min(max(max_len + 2, minimum), cap)
