"""Excel report for the signal-focused comparison.

Kept separate from the baseline report: the two answer different questions and
mixing their sheets into one workbook would invite reading a signal-contract
finding as a frame-level one.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from dbc_compare_tool.core.models import (
    SIGNAL_FOCUS_ATTENTION_STATUSES,
    SIGNAL_FOCUS_STATUSES,
    SignalFocusResult,
    SignalFocusRow,
    SignalRef,
)
from dbc_compare_tool.core.signal_focus import format_app_value, format_value_table
from dbc_compare_tool.report._style import (
    BORDER,
    autosize_columns,
    format_header_row,
)

# Ordered by how much application work the status implies.
_STATUS_FILL: dict[str, str] = {
    "Removed":           "FFC7CE",   # red — application code will break
    "Modified":          "FFF2CC",   # yellow — interpretation changed
    "Added":             "E2EFDA",   # green
    "Direction Changed": "FCE4D6",   # salmon — port direction flips
    "Out Of Node Scope": "FCE4D6",
    "Ambiguous":         "E4DFEC",   # purple — needs a human decision
    "Not In DBC":        "E4DFEC",
    "Moved":             "F2F2F2",   # gray — informational only
    "Unchanged":         "FFFFFF",
}

_VALUE_KIND_FILL: dict[str, str] = {
    "Relabeled":     "FFC7CE",   # same number, new meaning — the silent break
    "Value Removed": "FCE4D6",
    "Value Added":   "E2EFDA",
}

_OLD_VAL_FILL = "FCE4D6"
_NEW_VAL_FILL = "E2EFDA"


def write_signal_focus_report(result: SignalFocusResult, output_path: Path) -> Path:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Signal Focus Summary"
    signal_sheet = workbook.create_sheet("Signal Focus")
    property_sheet = workbook.create_sheet("Property Diff (App)")
    value_sheet = workbook.create_sheet("Value Table Diff")

    _write_summary(summary_sheet, result)
    _write_signals(signal_sheet, result)
    _write_property_diff(property_sheet, result)
    _write_value_table_diff(value_sheet, result)

    _format_summary(summary_sheet)
    _format_signal_sheet(signal_sheet)
    _format_property_sheet(property_sheet)
    _format_value_sheet(value_sheet)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


# ---------------------------------------------------------------------------
# Writers
# ---------------------------------------------------------------------------

def _write_summary(sheet, result: SignalFocusResult) -> None:
    generated = datetime.now().strftime("%Y-%m-%d %H:%M")
    sheet.append(["DBC Compare Tool  —  Signal Focus Report", generated])
    sheet.append([])

    sheet.append(["DBC File", "Old Node", "New Node"])
    for selection in result.selections:
        sheet.append([
            selection.dbc_file,
            selection.old_node or "—",
            selection.new_node or "—",
        ])
    sheet.append([])

    scope = (
        f"Signal list: {result.watchlist_size} signal(s)"
        if result.watchlist_size
        else "Signal list: none — full node audit"
    )
    sheet.append([scope])
    sheet.append([])

    sheet.append(["Status", "Count"])
    summary = result.summary()
    for status in SIGNAL_FOCUS_STATUSES:
        sheet.append([status, summary.get(status, 0)])
    sheet.append(["Needs Review", summary.get("Needs Review", 0)])
    sheet.append(["Total Signals", summary.get("Total Signals", 0)])


def _write_signals(sheet, result: SignalFocusResult) -> None:
    sheet.append([
        "Signal",
        "Status",
        "In Signal List",
        "Direction (Old)",
        "Direction (New)",
        "Length",
        "Value Type",
        "Factor",
        "Offset",
        "Min",
        "Max",
        "Unit",
        "Initial Value",
        "Value Table",
        "Carrier (Old)",
        "Carrier (New)",
        "Changed Properties",
        "Note",
    ])
    for row in result.rows:
        reference = _reference_signal(row)
        sheet.append([
            row.signal_name,
            row.status,
            "Yes" if row.in_watchlist else "No",
            _direction(row.old_refs),
            _direction(row.new_refs),
            _current(reference, "length"),
            _current(reference, "value_type"),
            _current(reference, "factor"),
            _current(reference, "offset"),
            _current(reference, "minimum"),
            _current(reference, "maximum"),
            _current(reference, "unit"),
            _current(reference, "raw_initial"),
            format_value_table(reference),
            _carriers(row.old_refs),
            _carriers(row.new_refs),
            _changed_properties(row),
            row.note,
        ])


def _write_property_diff(sheet, result: SignalFocusResult) -> None:
    sheet.append(["Signal", "Status", "In Signal List", "Property", "Old Value", "New Value"])
    for row in result.rows:
        for prop, old_value, new_value in row.property_diffs:
            sheet.append([
                row.signal_name,
                row.status,
                "Yes" if row.in_watchlist else "No",
                prop,
                old_value,
                new_value,
            ])


def _write_value_table_diff(sheet, result: SignalFocusResult) -> None:
    sheet.append(["Signal", "Raw Value", "Old Label", "New Label", "Kind"])
    for row in result.rows:
        for raw, old_label, new_label, kind in row.value_table_diffs:
            sheet.append([row.signal_name, raw, old_label or "—", new_label or "—", kind])


# ---------------------------------------------------------------------------
# Formatters
# ---------------------------------------------------------------------------

def _format_summary(sheet) -> None:
    title_cell = sheet.cell(row=1, column=1)
    title_cell.font = Font(bold=True, size=13, color="1F4E78")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[1].height = 26

    date_cell = sheet.cell(row=1, column=2)
    date_cell.font = Font(italic=True, size=10, color="595959")
    date_cell.alignment = Alignment(horizontal="right", vertical="center")

    for row in sheet.iter_rows(min_row=3):
        label = str(row[0].value or "")
        if label in ("DBC File", "Status"):
            format_header_row(sheet, row[0].row)
            continue
        if label in SIGNAL_FOCUS_STATUSES:
            fill = PatternFill("solid", fgColor=_STATUS_FILL.get(label, "FFFFFF"))
            for cell in row[:2]:
                cell.fill = fill
                cell.border = BORDER
        elif label in ("Needs Review", "Total Signals"):
            for cell in row[:2]:
                cell.font = Font(bold=True, size=11)
                cell.border = BORDER

    sheet.column_dimensions["A"].width = 42
    sheet.column_dimensions["B"].width = 22
    sheet.column_dimensions["C"].width = 22


def _format_signal_sheet(sheet) -> None:
    format_header_row(sheet)
    sheet.freeze_panes = "B2"
    sheet.auto_filter.ref = sheet.dimensions

    for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        status = str(row[1].value or "")
        fill = PatternFill("solid", fgColor=_STATUS_FILL.get(status, "FFFFFF"))
        for cell in row:
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = BORDER
        if status in SIGNAL_FOCUS_ATTENTION_STATUSES:
            row[0].font = Font(bold=True)
        sheet.row_dimensions[row_idx].height = 18

    autosize_columns(sheet, cap=55)


def _format_property_sheet(sheet) -> None:
    format_header_row(sheet)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = BORDER
        row[4].fill = PatternFill("solid", fgColor=_OLD_VAL_FILL)
        row[5].fill = PatternFill("solid", fgColor=_NEW_VAL_FILL)
        sheet.row_dimensions[row_idx].height = 18

    autosize_columns(sheet, cap=50)


def _format_value_sheet(sheet) -> None:
    format_header_row(sheet)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        kind = str(row[4].value or "")
        fill = PatternFill("solid", fgColor=_VALUE_KIND_FILL.get(kind, "FFFFFF"))
        for cell in row:
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = BORDER
        sheet.row_dimensions[row_idx].height = 18

    autosize_columns(sheet, cap=50)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _reference_signal(row: SignalFocusRow):
    """The signal whose values the overview columns show.

    The new baseline is what the application will be built against; a removed
    signal only exists on the old side, so fall back to it.
    """
    refs = row.new_refs or row.old_refs
    return refs[0].signal if refs else None


def _current(signal, attribute: str) -> str:
    if signal is None:
        return ""
    return format_app_value(getattr(signal, attribute))


def _direction(refs: tuple[SignalRef, ...]) -> str:
    if not refs:
        return "—"
    directions = {ref.direction for ref in refs}
    if directions == {"Tx"}:
        return "Tx"
    if directions == {"Rx"}:
        return "Rx"
    return "Tx/Rx"


def _carriers(refs: tuple[SignalRef, ...]) -> str:
    if not refs:
        return "—"
    seen: list[str] = []
    for ref in refs:
        label = f"{ref.message_name} (0x{ref.can_id:X})"
        if label not in seen:
            seen.append(label)
    return ", ".join(seen)


def _changed_properties(row: SignalFocusRow) -> str:
    parts = [f"{prop}: {old} -> {new}" for prop, old, new in row.property_diffs]
    parts += [
        f"Value {raw}: {old or '—'} -> {new or '—'} ({kind})"
        for raw, old, new, kind in row.value_table_diffs
    ]
    return "\n".join(parts)
