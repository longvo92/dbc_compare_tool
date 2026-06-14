from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from dbc_compare_tool.core.models import Change, ComparisonResult, FilePairSummary


SUMMARY_ORDER = [
    "Messages Added",
    "Messages Removed",
    "Messages Modified",
    "Messages Renamed",
    "Signals Added",
    "Signals Removed",
    "Signals Modified",
    "Signals Renamed",
    "Total Changes",
]

_HEADER_BG = "2E74B5"   # modern blue header
_HEADER_FG = "FFFFFF"

# Row background color keyed by change_type
_CHANGE_ROW_FILL: dict[str, str] = {
    "Added":           "E2EFDA",   # light green
    "Removed":         "FCE4D6",   # light salmon
    "Modified":        "FFF2CC",   # light yellow
    "Renamed":         "DDEBF7",   # light blue
    "Possible Rename": "EDF7FF",   # very light blue
}

# Confidence level cell highlight
_CONF_FILL: dict[str, str] = {
    "High":   "C6EFCE",   # green
    "Medium": "FFEB9C",   # yellow
    "Low":    "FFC7CE",   # pink-red
}

# Summary metric row colors (matches grouping)
_SUMMARY_METRIC_FILL: dict[str, str] = {
    "Messages Added":    "E8F8F5",
    "Messages Removed":  "FDEDEC",
    "Messages Modified": "FEF9E7",
    "Messages Renamed":  "EBF5FB",
    "Signals Added":     "E9F7EF",
    "Signals Removed":   "FDEDEC",
    "Signals Modified":  "FEF9E7",
    "Signals Renamed":   "EBF5FB",
    "Total Changes":     "F0F0F0",
}

_OLD_VAL_FILL = "FCE4D6"   # light salmon — what it was
_NEW_VAL_FILL = "E2EFDA"   # light green  — what it became

# Overview sheet status row colors
_STATUS_FILL: dict[str, str] = {
    "Matched":     "F2F2F2",   # light gray
    "DBC Added":   "E2EFDA",   # light green
    "DBC Removed": "FCE4D6",   # light salmon
    "DBC Renamed": "DDEBF7",   # light blue
}

_THIN = Side(style="thin", color="D0D0D0")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def write_excel_report(result: ComparisonResult, output_path: Path) -> Path:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    overview_sheet = workbook.create_sheet("DBC Overview")
    message_sheet = workbook.create_sheet("Message Details")
    signal_sheet = workbook.create_sheet("Signal Details")
    diff_sheet = workbook.create_sheet("Property Diff")

    _write_summary(summary_sheet, result)
    _write_overview(overview_sheet, result.file_pairs)
    _write_message_details(message_sheet, result.message_changes)
    _write_signal_details(signal_sheet, result.signal_changes)
    _write_property_diff(diff_sheet, result.message_changes, result.signal_changes)

    _format_summary(summary_sheet)
    _format_overview_sheet(overview_sheet)
    _format_detail_sheet(message_sheet, change_col=2, conf_score_col=6, conf_level_col=7)
    _format_detail_sheet(signal_sheet, change_col=3, conf_score_col=6, conf_level_col=7)
    _format_property_diff_sheet(diff_sheet)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


# ---------------------------------------------------------------------------
# Writers
# ---------------------------------------------------------------------------

def _write_summary(sheet, result: ComparisonResult) -> None:
    generated = datetime.now().strftime("%Y-%m-%d %H:%M")
    sheet.append(["DBC Compare Tool  —  Comparison Report", generated])
    sheet.append([])                                           # blank separator
    sheet.append(["Metric", "Count"])
    summary = result.summary()
    for metric in SUMMARY_ORDER:
        sheet.append([metric, summary[metric]])


def _write_overview(sheet, file_pairs: list[FilePairSummary]) -> None:
    sheet.append([
        "DBC File",
        "Status",
        "Old Path",
        "New Path",
        "Pairing Confidence",
        "Messages (Old)",
        "Messages (New)",
        "Signals (Old)",
        "Signals (New)",
    ])
    for fp in file_pairs:
        sheet.append([
            fp.dbc_file,
            fp.status,
            fp.old_path,
            fp.new_path,
            _fmt_confidence(fp.pairing_confidence),
            fp.message_count_old,
            fp.message_count_new,
            fp.signal_count_old,
            fp.signal_count_new,
        ])


def _write_message_details(sheet, changes: list[Change]) -> None:
    sheet.append([
        "DBC File",
        "Change Type",
        "Old Message Name",
        "New Message Name",
        "CAN ID",
        "Confidence Score",
        "Confidence Level",
        "Change Description",
    ])
    for change in changes:
        sheet.append([
            change.dbc_file,
            change.change_type,
            change.old_name,
            change.new_name,
            _fmt_can_id(change.can_id),
            _fmt_confidence(change.confidence),
            change.confidence_level,
            change.description,
        ])


def _write_signal_details(sheet, changes: list[Change]) -> None:
    sheet.append([
        "DBC File",
        "Parent Message",
        "Change Type",
        "Old Signal Name",
        "New Signal Name",
        "Confidence Score",
        "Confidence Level",
        "Changed Properties",
    ])
    for change in changes:
        sheet.append([
            change.dbc_file,
            change.parent_message,
            change.change_type,
            change.old_name,
            change.new_name,
            _fmt_confidence(change.confidence),
            change.confidence_level,
            change.description,
        ])


def _write_property_diff(
    sheet,
    message_changes: list[Change],
    signal_changes: list[Change],
) -> None:
    sheet.append([
        "DBC File",
        "Entity Type",
        "Old Name",
        "New Name",
        "Parent Message",
        "Change Type",
        "Property",
        "Old Value",
        "New Value",
    ])
    for change in message_changes:
        for prop, old_val, new_val in change.property_diffs:
            sheet.append([
                change.dbc_file,
                "Message",
                change.old_name,
                change.new_name,
                "",
                change.change_type,
                prop,
                old_val,
                new_val,
            ])
    for change in signal_changes:
        for prop, old_val, new_val in change.property_diffs:
            sheet.append([
                change.dbc_file,
                "Signal",
                change.old_name,
                change.new_name,
                change.parent_message,
                change.change_type,
                prop,
                old_val,
                new_val,
            ])


# ---------------------------------------------------------------------------
# Formatters
# ---------------------------------------------------------------------------

def _format_summary(sheet) -> None:
    # Row 1 — report title
    title_cell = sheet.cell(row=1, column=1)
    title_cell.font = Font(bold=True, size=13, color="1F4E78")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[1].height = 26

    date_cell = sheet.cell(row=1, column=2)
    date_cell.font = Font(italic=True, size=10, color="595959")
    date_cell.alignment = Alignment(horizontal="right", vertical="center")

    # Row 3 — column headers
    header_fill = PatternFill("solid", fgColor=_HEADER_BG)
    header_font = Font(color=_HEADER_FG, bold=True, size=11)
    for cell in sheet[3]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _BORDER
    sheet.row_dimensions[3].height = 22

    # Rows 4+ — metric data
    for row_idx, metric in enumerate(SUMMARY_ORDER, start=4):
        row = sheet[row_idx]
        bg = _SUMMARY_METRIC_FILL.get(metric, "FFFFFF")
        fill = PatternFill("solid", fgColor=bg)
        for cell in row:
            cell.fill = fill
            cell.border = _BORDER
            cell.alignment = Alignment(vertical="center")
        if metric == "Total Changes":
            for cell in row:
                cell.font = Font(bold=True, size=11)
        sheet.row_dimensions[row_idx].height = 18

    sheet.column_dimensions["A"].width = 26
    sheet.column_dimensions["B"].width = 10
    sheet.freeze_panes = "A4"


def _format_overview_sheet(sheet) -> None:
    header_fill = PatternFill("solid", fgColor=_HEADER_BG)
    header_font = Font(color=_HEADER_FG, bold=True, size=11)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER
    sheet.row_dimensions[1].height = 26

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        status = str(row[1].value or "")
        row_fill = PatternFill("solid", fgColor=_STATUS_FILL.get(status, "FFFFFF"))
        for cell in row:
            cell.fill = row_fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = _BORDER
        sheet.row_dimensions[row_idx].height = 18

    for col_cells in sheet.columns:
        max_len = max(len(str(cell.value or "")) for cell in col_cells)
        col_letter = get_column_letter(col_cells[0].column)
        sheet.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 60)


def _format_detail_sheet(
    sheet,
    change_col: int,
    conf_score_col: int,
    conf_level_col: int,
) -> None:
    header_fill = PatternFill("solid", fgColor=_HEADER_BG)
    header_font = Font(color=_HEADER_FG, bold=True, size=11)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER
    sheet.row_dimensions[1].height = 26

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        change_type = str(row[change_col - 1].value or "")
        row_bg = _CHANGE_ROW_FILL.get(change_type, "FFFFFF")
        row_fill = PatternFill("solid", fgColor=row_bg)

        for cell in row:
            cell.fill = row_fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = _BORDER

        # Highlight confidence level cell individually
        conf_cell = row[conf_level_col - 1]
        conf_level = str(conf_cell.value or "")
        if conf_level in _CONF_FILL:
            conf_cell.fill = PatternFill("solid", fgColor=_CONF_FILL[conf_level])
            conf_cell.font = Font(bold=True)
            conf_cell.alignment = Alignment(horizontal="center", vertical="top")

        # Center-align confidence score
        row[conf_score_col - 1].alignment = Alignment(horizontal="center", vertical="top")

        sheet.row_dimensions[row_idx].height = 18

    # Auto-width columns (capped at 60)
    for col_cells in sheet.columns:
        max_len = max(len(str(cell.value or "")) for cell in col_cells)
        col_letter = get_column_letter(col_cells[0].column)
        sheet.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 60)


def _format_property_diff_sheet(sheet) -> None:
    header_fill = PatternFill("solid", fgColor=_HEADER_BG)
    header_font = Font(color=_HEADER_FG, bold=True, size=11)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER
    sheet.row_dimensions[1].height = 26

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        change_type = str(row[5].value or "")  # column 6 = Change Type
        row_fill = PatternFill("solid", fgColor=_CHANGE_ROW_FILL.get(change_type, "FFFFFF"))

        for cell in row:
            cell.fill = row_fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = _BORDER

        row[7].fill = PatternFill("solid", fgColor=_OLD_VAL_FILL)  # Old Value — salmon
        row[8].fill = PatternFill("solid", fgColor=_NEW_VAL_FILL)  # New Value — green
        sheet.row_dimensions[row_idx].height = 18

    for col_cells in sheet.columns:
        max_len = max(len(str(cell.value or "")) for cell in col_cells)
        col_letter = get_column_letter(col_cells[0].column)
        sheet.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 50)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _fmt_confidence(confidence: float | None) -> str:
    return "" if confidence is None else f"{confidence:.2f}"


def _fmt_can_id(can_id: int | None) -> str:
    if can_id is None:
        return ""
    try:
        return f"0x{int(can_id):X}"
    except (TypeError, ValueError):
        return str(can_id)
